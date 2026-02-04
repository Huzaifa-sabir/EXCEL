from flask import Flask, request, send_file, render_template_string
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from io import BytesIO
import json
import traceback

app = Flask(__name__)

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Visa Excel Generator</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }

        .container {
            max-width: 900px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            overflow: hidden;
        }

        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }

        .header h1 {
            font-size: 2.5rem;
            margin-bottom: 10px;
        }

        .header p {
            opacity: 0.9;
            font-size: 1.1rem;
        }

        .form-container {
            padding: 40px;
        }

        .form-group {
            margin-bottom: 25px;
        }

        .form-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
        }

        label {
            display: block;
            margin-bottom: 8px;
            color: #333;
            font-weight: 600;
            font-size: 0.95rem;
        }

        input, select {
            width: 100%;
            padding: 12px 16px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 1rem;
            transition: all 0.3s ease;
        }

        input:focus, select:focus {
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }

        .add-button, .download-button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 14px 28px;
            border-radius: 8px;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
            margin-right: 10px;
        }

        .add-button:hover, .download-button:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }

        .add-button:active, .download-button:active {
            transform: translateY(0);
        }

        .download-button:disabled {
            opacity: 0.6;
            cursor: not-allowed;
            transform: none;
        }

        .customers-list {
            margin-top: 40px;
            padding-top: 40px;
            border-top: 2px solid #f0f0f0;
        }

        .customers-list h2 {
            color: #333;
            margin-bottom: 20px;
            font-size: 1.8rem;
        }

        .customer-card {
            background: #f8f9fa;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 15px;
            border-left: 4px solid #667eea;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .customer-info {
            flex: 1;
        }

        .customer-name {
            font-weight: 600;
            color: #333;
            font-size: 1.1rem;
            margin-bottom: 5px;
        }

        .customer-details {
            color: #666;
            font-size: 0.9rem;
        }

        .remove-button {
            background: #ff4757;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            transition: background 0.2s ease;
        }

        .remove-button:hover {
            background: #ff3838;
        }

        .button-group {
            margin-top: 30px;
            display: flex;
            gap: 10px;
        }

        .empty-state {
            text-align: center;
            padding: 40px;
            color: #999;
        }

        .required {
            color: #ff4757;
        }

        @media (max-width: 768px) {
            .form-row {
                grid-template-columns: 1fr;
            }
            .header h1 {
                font-size: 1.8rem;
            }
            .form-container {
                padding: 20px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🛂 Visa Excel Generator</h1>
            <p>Create formatted Excel files with dynamic dropdowns</p>
        </div>
        
        <div class="form-container">
            <h2 style="margin-bottom: 20px; color: #333">Customer Details</h2>
            
            <form id="customerForm">
                <div class="form-row">
                    <div class="form-group">
                        <label>City</label>
                        <select name="city" id="city">
                            <option value="Bissau">Bissau</option>
                        </select>
                    </div>
                    
                    <div class="form-group">
                        <label>Category</label>
                        <select name="category" id="category">
                            <option value="National Visa">National Visa</option>
                            <option value="Schengen Visa">Schengen Visa</option>
                        </select>
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Subcategory</label>
                        <select name="subcategory" id="subcategory">
                            <option value="Work Visa">Work Visa</option>
                        </select>
                    </div>
                    
                    <div class="form-group">
                        <label>Price</label>
                        <input type="number" name="price" id="price" placeholder="Enter price">
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Last Name <span class="required">*</span></label>
                        <input type="text" name="lastName" id="lastName" placeholder="Enter last name" required>
                    </div>
                    
                    <div class="form-group">
                        <label>First Name <span class="required">*</span></label>
                        <input type="text" name="firstName" id="firstName" placeholder="Enter first name" required>
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Passport Number</label>
                        <input type="text" name="passportNumber" id="passportNumber" placeholder="C00000000">
                    </div>
                    
                    <div class="form-group">
                        <label>Birthdate (dd.mm.yyyy)</label>
                        <input type="text" name="birthdate" id="birthdate" placeholder="13.03.1994">
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Passport Validity (dd.mm.yyyy)</label>
                        <input type="text" name="passportValidity" id="passportValidity" placeholder="12.11.2029">
                    </div>
                    
                    <div class="form-group">
                        <label>Gender</label>
                        <select name="gender" id="gender">
                            <option value="Male">Male</option>
                            <option value="Female">Female</option>
                        </select>
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Phone (with country code)</label>
                        <input type="text" name="phone" id="phone" placeholder="245857456140">
                    </div>
                    
                    <div class="form-group">
                        <label>Nationality</label>
                        <select name="nationality" id="nationality">
                            <option value="GUINEA-BISSAU">GUINEA-BISSAU</option>
                            <option value="SENEGAL">SENEGAL</option>
                        </select>
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Book Date From (dd.mm.yyyy)</label>
                        <input type="text" name="bookDateFrom" id="bookDateFrom" placeholder="01.01.2025">
                    </div>
                    
                    <div class="form-group">
                        <label>Book Date To (dd.mm.yyyy)</label>
                        <input type="text" name="bookDateTo" id="bookDateTo" placeholder="31.01.2025">
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Agent Name <span class="required">*</span></label>
                        <input type="text" name="agentName" id="agentName" placeholder="Enter agent name" required>
                    </div>
                    
                    <div class="form-group">
                        <label>Days Gap</label>
                        <input type="number" name="daysGap" id="daysGap" placeholder="Enter days gap">
                    </div>
                </div>
                
                <div class="form-row">
                    <div class="form-group">
                        <label>Group (not required)</label>
                        <input type="text" name="group" id="group" placeholder="Enter group">
                    </div>
                    
                    <div class="form-group">
                        <label>Email</label>
                        <input type="email" name="email" id="email" placeholder="example@email.com">
                    </div>
                </div>
                
                <div class="button-group">
                    <button type="button" class="add-button" onclick="addCustomer()">➕ Add Customer</button>
                </div>
            </form>
            
            <div id="customersList" class="customers-list" style="display: none;">
                <h2>Added Customers (<span id="customerCount">0</span>)</h2>
                <div id="customersContainer"></div>
                
                <div class="button-group">
                    <button class="download-button" onclick="generateExcel()" id="downloadBtn">
                        📥 Download Excel File
                    </button>
                </div>
            </div>
            
            <div id="emptyState" class="empty-state">
                <p>No customers added yet. Add your first customer above.</p>
            </div>
        </div>
    </div>

    <script>
        let customers = [];
        
        const subcategoryOptions = {
            'National Visa': ['Work Visa', 'Job Seeker Visa', 'Medical Treatment Visa', 'Study Visa', 'Family Reunion Visa'],
            'Schengen Visa': ['Schengen Visa']
        };
        
        document.getElementById('category').addEventListener('change', function() {
            const category = this.value;
            const subcategorySelect = document.getElementById('subcategory');
            subcategorySelect.innerHTML = '';
            
            subcategoryOptions[category].forEach(option => {
                const opt = document.createElement('option');
                opt.value = option;
                opt.textContent = option;
                subcategorySelect.appendChild(opt);
            });
        });
        
        function addCustomer() {
            const form = document.getElementById('customerForm');
            
            if (!form.checkValidity()) {
                alert('Please fill in all required fields');
                return;
            }
            
            const customer = {
                city: document.getElementById('city').value,
                category: document.getElementById('category').value,
                subcategory: document.getElementById('subcategory').value,
                price: document.getElementById('price').value,
                lastName: document.getElementById('lastName').value,
                firstName: document.getElementById('firstName').value,
                passportNumber: document.getElementById('passportNumber').value,
                birthdate: document.getElementById('birthdate').value,
                passportValidity: document.getElementById('passportValidity').value,
                gender: document.getElementById('gender').value,
                phone: document.getElementById('phone').value,
                nationality: document.getElementById('nationality').value,
                bookDateFrom: document.getElementById('bookDateFrom').value,
                bookDateTo: document.getElementById('bookDateTo').value,
                agentName: document.getElementById('agentName').value,
                daysGap: document.getElementById('daysGap').value,
                group: document.getElementById('group').value,
                email: document.getElementById('email').value
            };
            
            customers.push(customer);
            updateCustomersList();
            
            // Reset form
            document.getElementById('lastName').value = '';
            document.getElementById('firstName').value = '';
            document.getElementById('passportNumber').value = '';
            document.getElementById('birthdate').value = '';
            document.getElementById('passportValidity').value = '';
            document.getElementById('phone').value = '';
            document.getElementById('bookDateFrom').value = '';
            document.getElementById('bookDateTo').value = '';
            document.getElementById('agentName').value = '';
            document.getElementById('daysGap').value = '';
            document.getElementById('group').value = '';
            document.getElementById('email').value = '';
            document.getElementById('price').value = '';
        }
        
        function removeCustomer(index) {
            customers.splice(index, 1);
            updateCustomersList();
        }
        
        function updateCustomersList() {
            const container = document.getElementById('customersContainer');
            const count = document.getElementById('customerCount');
            const listDiv = document.getElementById('customersList');
            const emptyState = document.getElementById('emptyState');
            
            count.textContent = customers.length;
            
            if (customers.length > 0) {
                listDiv.style.display = 'block';
                emptyState.style.display = 'none';
                
                container.innerHTML = customers.map((customer, index) => `
                    <div class="customer-card">
                        <div class="customer-info">
                            <div class="customer-name">${customer.firstName} ${customer.lastName}</div>
                            <div class="customer-details">${customer.category} • ${customer.subcategory} • Agent: ${customer.agentName}</div>
                        </div>
                        <button class="remove-button" onclick="removeCustomer(${index})">Remove</button>
                    </div>
                `).join('');
            } else {
                listDiv.style.display = 'none';
                emptyState.style.display = 'block';
            }
        }
        
        async function generateExcel() {
            if (customers.length === 0) {
                alert('Please add at least one customer');
                return;
            }
            
            const btn = document.getElementById('downloadBtn');
            btn.disabled = true;
            btn.textContent = '⏳ Generating...';
            
            try {
                const response = await fetch('/generate', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({ customers })
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `GUINEA_VISA_${new Date().toISOString().split('T')[0]}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    alert('✅ Excel file generated successfully!');
                } else {
                    alert('❌ Error generating Excel file');
                }
            } catch (error) {
                console.error('Error:', error);
                alert('❌ Error: ' + error.message);
            } finally {
                btn.disabled = false;
                btn.textContent = '📥 Download Excel File';
            }
        }
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        customers = data.get('customers', [])
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main Data sheet
    data_sheet = wb.create_sheet('Data', 0)
    
    # Create hidden sheets
    sheet2 = wb.create_sheet('Лист2', 1)
    sheet2.sheet_state = 'hidden'
    
    sheet3 = wb.create_sheet('Лист3', 2)
    sheet3.sheet_state = 'hidden'
    
    # Setup Лист2 (Categories and Subcategories)
    sheet2['A1'] = 'City'
    sheet2['B1'] = 'Category'
    sheet2['C1'] = 'National Visa Subcategory'
    sheet2['D1'] = 'Schengen Visa Subcategory'
    
    sheet2['A2'] = 'Bissau'
    sheet2['B2'] = 'National Visa'
    sheet2['C2'] = 'Work Visa'
    sheet2['D2'] = 'Schengen Visa'
    
    sheet2['B3'] = 'Schengen Visa'
    sheet2['C3'] = 'Job Seeker Visa'
    
    sheet2['C4'] = 'Medical Treatment Visa'
    sheet2['C5'] = 'Study Visa'
    sheet2['C6'] = 'Family Reunion Visa'
    
    # Define named ranges for Лист2
    wb.defined_names['City'] = DefinedName('City', attr_text="'Лист2'!$A$2")
    wb.defined_names['Category'] = DefinedName('Category', attr_text="'Лист2'!$B$2:$B$3")
    wb.defined_names['National_Visa_Subcategory'] = DefinedName('National_Visa_Subcategory', attr_text="'Лист2'!$C$2:$C$6")
    wb.defined_names['Schengen_Visa_Subcategory'] = DefinedName('Schengen_Visa_Subcategory', attr_text="'Лист2'!$D$2")
    
    # Setup Лист3 (Gender and Country)
    sheet3['A1'] = 'Gender'
    sheet3['B1'] = 'Country'
    sheet3['A2'] = 'Male'
    sheet3['B2'] = 'GUINEA-BISSAU'
    sheet3['A3'] = 'Female'
    sheet3['B3'] = 'SENEGAL'
    
    # Define named ranges for Лист3
    wb.defined_names['Gender'] = DefinedName('Gender', attr_text="'Лист3'!$A$2:$A$3")
    wb.defined_names['Country'] = DefinedName('Country', attr_text="'Лист3'!$B$2:$B$3")
    
    # Setup Data sheet headers
    headers = [
        'City', 'Category', 'Subcategory', 'Price', 'Last Name', 'First Name',
        'Passport number', 'Birthdate (dd.mm.yyyy)', 'Passport validity  (dd.mm.yyyy)',
        'Gender (M/F)', 'Phone (with country code)', 'Nationality',
        'Book date from  (dd.mm.yyyy)', 'Book date to  (dd.mm.yyyy)',
        'Agent Name (required)', 'Days gap', 'Group (not required)', 'e-mail'
    ]
    
    data_sheet.append(headers)
    
    # Set column widths
    column_widths = [9.57, 16.14, 18.71, 5.14, 15.71, 25.43, 15.43, 12.14, 13.14, 
                     6.57, 23.86, 9.43, 12.43, 11.86, 16.5, 5.43, 10.43, 24.43]
    
    for idx, width in enumerate(column_widths, 1):
        data_sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Set row heights
    data_sheet.row_dimensions[1].height = 33.0
    
    # Add customer data
    for customer in customers:
        row = [
            customer.get('city', ''),
            customer.get('category', ''),
            customer.get('subcategory', ''),
            customer.get('price', ''),
            customer.get('lastName', ''),
            customer.get('firstName', ''),
            customer.get('passportNumber', ''),
            customer.get('birthdate', ''),
            customer.get('passportValidity', ''),
            customer.get('gender', ''),
            customer.get('phone', ''),
            customer.get('nationality', ''),
            customer.get('bookDateFrom', ''),
            customer.get('bookDateTo', ''),
            customer.get('agentName', ''),
            customer.get('daysGap', ''),
            customer.get('group', ''),
            customer.get('email', '')
        ]
        data_sheet.append(row)
        data_sheet.row_dimensions[data_sheet.max_row].height = 14.25
    
    # Add data validations (dropdowns) for 200 rows
    total_rows = 202
    
    # City dropdown (A2:A202)
    dv_city = DataValidation(type="list", formula1='City', allow_blank=True)
    data_sheet.add_data_validation(dv_city)
    dv_city.add(f'A2:A{total_rows}')
    
    # Category dropdown (B2:B202)
    dv_category = DataValidation(type="list", formula1='Category', allow_blank=True)
    data_sheet.add_data_validation(dv_category)
    dv_category.add(f'B2:B{total_rows}')
    
    # Subcategory dropdown with INDIRECT formula (C2:C202)
    dv_subcategory = DataValidation(
        type="list",
        formula1='INDIRECT(SUBSTITUTE(SUBSTITUTE(B2," ","_"),"-","_") & "_Subcategory")',
        allow_blank=True,
        showErrorMessage=True
    )
    data_sheet.add_data_validation(dv_subcategory)
    dv_subcategory.add(f'C2:C{total_rows}')
    
    # Price validation (D2:D202)
    dv_price = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=1, allow_blank=True)
    data_sheet.add_data_validation(dv_price)
    dv_price.add(f'D2:D{total_rows}')
    
    # Gender dropdown (J2:J202)
    dv_gender = DataValidation(type="list", formula1='Gender', allow_blank=True)
    data_sheet.add_data_validation(dv_gender)
    dv_gender.add(f'J2:J{total_rows}')
    
    # Nationality dropdown (L2:L202)
    dv_nationality = DataValidation(type="list", formula1='Country', allow_blank=True)
    data_sheet.add_data_validation(dv_nationality)
    dv_nationality.add(f'L2:L{total_rows}')
    
    # Days gap validation (P2:P202)
    dv_days = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=1, allow_blank=True)
    data_sheet.add_data_validation(dv_days)
    dv_days.add(f'P2:P{total_rows}')
    
    # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'GUINEA_VISA.xlsx'
        )
    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        print(traceback.format_exc())
        return {'error': str(e), 'traceback': traceback.format_exc()}, 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
